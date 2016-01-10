
import os
from openpyxl import load_workbook

def isFile(fn):
	def wrapper(*args,**kwargs):
		if os.path.isfile(args[1]):
			pass
		else:
			print "File Does'nt exists"	
	return wrapper
			
			
		
		

	

class xlsxreader():
	
	def read(self,path):
		files=[]
		flag=1
		try:
			workbook = load_workbook(path, use_iterators=True)
			first_sheet = workbook.get_sheet_names()[0]
			worksheet = workbook.get_sheet_by_name(first_sheet)
		except:
			flag=0
			print "Exception Here File Not Found or First Sheet Missing"
		try:
			for row in worksheet.iter_rows():
				try:
					Ticker=row[0].value
					Company_name=row[1].value
					erev=row[2].value
					Base_url=row[3].value
					if Ticker==None or Company_name==None or Base_url==None or Ticker=="Ticker":
						pass
					else:
						files.append((Ticker,Company_name,erev,Base_url))
				except:
					flag=0
					print "Exception in reading worksheet rows"
		except:
			print "can't read local files"
			files=[]
		return files
		

print xlsxreader().read("company.xlsx")
		

		
